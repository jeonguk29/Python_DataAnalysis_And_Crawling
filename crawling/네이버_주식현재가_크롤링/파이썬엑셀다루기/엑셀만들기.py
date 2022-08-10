import openpyxl

# 1) 액셀 만들기
wb = openpyxl.Workbook()

# 2) 엑셀 워크 시트 만들기 
ws = wb.create_sheet('오징어게임')


# 3) 데이터 추가하기  
ws['A1']  = "참가번호" # A1 셀에 값 입력
ws['B1']  = "성명"


ws['A2'] = 1
ws['B2'] = "오일남"

# 4) 엑셀 저장하기
wb.save(r'C:\포폴\GIT_PULL\GIT_Python_Data_Analysis_Crawling\crawling\네이버_주식현재가_크롤링\파이썬엑셀다루기\참가자_data.xlsx') # 저장할이름 과 확장자 넣어주면됨 
# C:\포폴\GIT_PULL\GIT_Python_Data_Analysis_Crawling\crawling\네이버_주식현재가_크롤링\파이썬엑셀다루기   원하는 경로 까지 넣어주면 거기로 저장됨 

# 앞에 r 옾션을 줘야지만 문자형태로 해서 경로 가능함 \ 이게 원화 이스케이프 문자로 바뀌는걸 방지함 